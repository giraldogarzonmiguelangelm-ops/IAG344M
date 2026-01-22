import re
from openpyxl import load_workbook
# ==========================
#  FUNCIÓN clean_id
#  ELIMINA CARACTERES NO NUMERICOS DE UN DOCUMENTO
#  "cc75.888.56" = "7588856"
# ==========================
def clean_id(value):
    # Elimina caracteres no numericos de un documento
    if value is None:
        return ""
    return re.sub(r"\D","",str(value))  #str sirve para transformar los digitos en cadena de texto para que tome el valor como string
# ==========================
#  FUNCIÓN merge_name
#  Une nombre y apellido en un solo campo
# ==========================
def merge_name(first_name, last_name):
    if first_name is None:
        first_name = ""
    if last_name is None:
        last_name = ""
    return f"{first_name} {last_name}".strip() #El strip elimina los espacios en blanco al inicio y al final
# ==========================
# FUNCIÓN PARA PROCESAR EL ARCHIVO DE EXCEL
# ==========================

def process_excel(path):
    #Acceso a la hoja llamada "Datos"
    wb = load_workbook(path)  #Carga el libro de excel
    ws = wb["Datos"]          #Accede a la hoja llamada Datos de Excel
    #Recorrer todas las filas desde la fila 2 
    for row in range (2,ws.max_row+1):    #Hace un recorrido desde la fila 2 hasta la ultima fila con el +1
        #Columna D: Identificador limpio
        ws[f"D{row}"] = clean_id(ws[f"A{row}"].value) #Guarda en la columna D el valor de la columna A pero aplicandole la funcion, que limpia el id
        #Columna E: Nombre Completo
        ws[f"E{row}"] = merge_name(
        ws[f"B{row}"].value, 
        ws[f"C{row}"].value
        )
        wb.save(path) #Guarda los cambios del archivo de Excel
# ==========================
#  FUNCIÓN DE COMPROBACIÓN(CONTROLADOR DE ERRORES)
# ==========================
def process_excel_safe(path):
    try:
        process_excel(path)
        return True, "Archivo procesado correctamente."
    except PermissionError: #Error cuando se intenta ejecutar con el archivo abierto
        return(
            False,
            "El archivo Excel esta abierto.\n"
            "Por favor, ciérrelo e intente de nuevo."
        )
    except KeyError:  #Error si no encuentra la hoja de Excel
        return False, "Hoja 'Datos' no encontrada"
    except Exception as e:  #Errores no conocidos con la variable temporal e
        return False, f"Error inesperado: {str(e)}"
        

